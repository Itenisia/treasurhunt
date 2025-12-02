import csv
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction

from armies.models import Faction, UnitType


class Command(BaseCommand):
    help = "Importe les unités depuis war3_units.csv ou war3_units_full.xlsx et crée les factions RSC/ESD/Chimie/BIO"

    def handle(self, *args, **options):
        base = Path(__file__).resolve().parent.parent.parent.parent
        xlsx_path = base / "war3_units_full.xlsx"
        csv_path = base / "war3_units.csv"

        mapping = {
            "Human": ("RSC", "RSC"),
            "Orc": ("ESD", "ESD"),
            "Undead": ("Chimie", "Chimie"),
            "NightElf": ("BIO", "BIO"),
            "Night Elf": ("BIO", "BIO"),
        }

        factions = {}
        for name, code in mapping.values():
            factions[code], _ = Faction.objects.get_or_create(code=code, defaults={"name": name})

        created, updated = 0, 0

        def upsert_unit(row, src_faction):
            nonlocal created, updated
            fac_code = mapping.get(src_faction, (None, None))[0]
            faction = factions.get(fac_code)
            gold = int(row.get("Gold") or 0)
            wood = int(row.get("Wood") or 0)
            cost = gold + wood
            damage_min = float(row.get("MinDamage") or 0)
            damage_max = float(row.get("MaxDamage") or damage_min or 1)
            defense = int(row.get("Armor") or 0)
            health = int(row.get("HP") or 1)
            rng_raw = row.get("Range")
            try:
                rng = float(rng_raw)
            except (TypeError, ValueError):
                rng = 600 if str(row.get("RangeType") or "").lower().startswith("rang") else 100
            range_cells = max(1, int(round(rng / 100)))
            aps_raw = row.get("AttacksPerSecond")
            try:
                aps = float(aps_raw)
            except (TypeError, ValueError):
                aps = 1.5
            attack_type = str(row.get("AttackType") or "normal").strip().lower() or "normal"
            armor_type = str(row.get("ArmorType") or "unarmored").strip().lower() or "unarmored"
            pop_cost = int(row.get("Pop") or 1)
            name = row["Unit"]
            ut, was_created = UnitType.objects.update_or_create(
                name=name,
                defaults={
                    "faction": faction,
                    "description": f"Import {src_faction}",
                    "cost": cost,
                    "defense": defense,
                    "health": health,
                    "attack_speed": min(4.0, aps),
                    "move_speed": 1.0,
                    "range": range_cells,
                    "damage_min": round(damage_min, 2),
                    "damage_max": round(damage_max, 2),
                    "crit_chance": 0.1,
                    "crit_multiplier": 2.0,
                    "dodge_chance": 0.0,
                    "aoe_radius": 0,
                    "speed": 1,
                    "attack_type": attack_type,
                    "armor_type": armor_type,
                    "pop_cost": pop_cost,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        with transaction.atomic():
            if xlsx_path.exists():
                from openpyxl import load_workbook

                wb = load_workbook(xlsx_path, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    rows_iter = ws.iter_rows(values_only=True)
                    headers = [str(c).strip() if c else "" for c in next(rows_iter)]
                    for row in rows_iter:
                        data = dict(zip(headers, row))
                        if not data.get("Unit"):
                            continue
                        upsert_unit(data, sheet)
            elif csv_path.exists():
                with csv_path.open() as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        upsert_unit(row, row.get("Faction"))
            else:
                self.stderr.write("Aucun fichier war3_units.csv ou war3_units_full.xlsx trouvé.")
                return

        self.stdout.write(self.style.SUCCESS(f"Import terminé. Créés: {created}, mis à jour: {updated}"))
