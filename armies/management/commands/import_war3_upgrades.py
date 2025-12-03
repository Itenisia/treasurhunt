import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from armies.models import Faction, UnitType, Upgrade


class Command(BaseCommand):
    help = "Importe les upgrades (niveau 1) depuis war3_upgrades_full_structured.xlsx en fusionnant or+bois."

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - guardrail pour env incomplet
            raise CommandError("openpyxl manquant. Installez les dépendances avec `pip install -r requirements.txt`.") from exc

        base = Path(__file__).resolve().parent.parent.parent.parent
        xlsx_path = base / "war3_upgrades_full_structured.xlsx"
        if not xlsx_path.exists():
            raise CommandError("Fichier war3_upgrades_full_structured.xlsx introuvable à la racine du projet.")

        mapping = {
            "Human": ("RSC", "RSC"),
            "Orc": ("ESD", "ESD"),
            "Undead": ("Chimie", "Chimie"),
            "NightElf": ("BIO", "BIO"),
            "Night Elf": ("BIO", "BIO"),
        }

        factions = {}
        for name, code in mapping.values():
            factions[code], _ = Faction.objects.get_or_create(code=code, defaults={"name": name})

        wb = load_workbook(xlsx_path, data_only=True)

        created = 0
        updated = 0
        skipped = 0
        missing_units = []

        def to_int(value):
            try:
                return int(float(value or 0))
            except (TypeError, ValueError):
                return 0

        def detect_bonus(effect: str) -> tuple[int, int]:
            lower = effect.lower()
            attack_bonus = 0
            defense_bonus = 1 if "armor" in lower else 0
            return attack_bonus, defense_bonus

        def parse_attack_pct(note: str) -> float:
            m = re.search(r"([0-9]+)", note or "")
            if not m:
                return 0.0
            return int(m.group(1)) / 100.0

        alias_map = {
            "wind rider": "Wyvern Rider",
            "demolisher": "Demolisher",
            "statue": "Obsidian Statue",
        }
        ignore_units = {
            "human buildings",
            "orc buildings",
            "ancients",
            "tree of life line",
            "burrow",
            "watch tower",
            "etc.",
        }

        @transaction.atomic
        def import_sheet(sheet_name: str):
            nonlocal created, updated, skipped, missing_units

            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers = [str(c).strip() if c else "" for c in next(rows_iter)]
            except StopIteration:
                return
            header_map = {h: i for i, h in enumerate(headers)}

            def val(row, key):
                idx = header_map.get(key)
                return row[idx] if idx is not None and idx < len(row) else None

            for row in rows_iter:
                level = to_int(val(row, "Level"))
                if level != 1:
                    continue
                effect = str(val(row, "Effect") or "").strip()
                name = str(val(row, "Upgrade") or "").strip()
                gold = to_int(val(row, "Gold"))
                wood = to_int(val(row, "Wood"))
                units_raw = str(val(row, "UnitsAffected") or "").strip()
                note = str(val(row, "DPS_Expected_Note") or "").strip()
                cost = gold + wood

                attack_bonus, defense_bonus = detect_bonus(effect)
                attack_pct = parse_attack_pct(note if "dps" in note.lower() else effect)
                if attack_bonus == 0 and attack_pct == 0 and defense_bonus == 0:
                    skipped += 1
                    continue

                desc_parts = [effect] if effect else []
                if units_raw:
                    desc_parts.append(f"Applique à: {units_raw}")
                if note:
                    desc_parts.append(f"Note: {note}")
                description = " | ".join(desc_parts)

                defaults = {
                    "faction": factions.get(mapping.get(sheet_name, (None, None))[0]),
                    "description": description,
                    "cost": cost,
                    "attack_bonus": attack_bonus,
                    "attack_bonus_pct": attack_pct,
                    "defense_bonus": defense_bonus,
                    "health_bonus": 0,
                    "speed_bonus": 0,
                    "unit_type": None,
                }
                upgrade, was_created = Upgrade.objects.update_or_create(
                    name=name,
                    defaults=defaults,
                )

                unit_names = [u.strip() for u in units_raw.split(",") if u and u.strip()]
                targets = []
                for uname in unit_names:
                    base = uname.strip()
                    key = base.lower()
                    if key in ignore_units or sheet_name == "Human":
                        continue
                    mapped = alias_map.get(key, base)
                    qs = UnitType.objects.filter(name__iexact=uname)
                    if not qs.exists() and mapped != base:
                        qs = UnitType.objects.filter(name__iexact=mapped)
                    if not qs.exists() and "/" in uname:
                        # Essaye chaque variante séparée par / (ex: Headhunter/Berserker)
                        for part in re.split(r"[\\/]", uname):
                            part = part.strip()
                            if not part:
                                continue
                            qs = UnitType.objects.filter(name__iexact=part)
                            if qs.exists():
                                break
                    if qs.exists():
                        targets.extend(list(qs))
                    else:
                        missing_units.append((name, uname))

                upgrade.unit_types.set(targets)

                if was_created:
                    created += 1
                else:
                    updated += 1

        for sheet in wb.sheetnames:
            if sheet not in mapping:
                continue
            import_sheet(sheet)

        summary = f"Import terminé. Créés: {created}, mis à jour: {updated}, ignorés (hors attaque/armure): {skipped}"
        self.stdout.write(self.style.SUCCESS(summary))
        if missing_units:
            seen = {}
            warn_lines = []
            for upg, unit in missing_units:
                key = (upg, unit)
                if key in seen:
                    continue
                seen[key] = True
                warn_lines.append(f" - {upg}: unité introuvable '{unit}'")
            self.stdout.write(self.style.WARNING("Unités non trouvées:\n" + "\n".join(warn_lines)))
