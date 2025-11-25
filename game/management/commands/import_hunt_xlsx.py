import os

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from game.models import Hunt, Step


class Command(BaseCommand):
    help = (
        "Importe une chasse au trésor depuis un fichier XLSX.\n"
        "Ligne 1 : titre | description | indice de départ\n"
        "Lignes suivantes : ordre | nom de l'étape | indice suivant"
    )

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Chemin vers le fichier XLSX à importer")

    def handle(self, *args, **options):
        path = options["xlsx_path"]
        if not os.path.isfile(path):
            raise CommandError(f"Fichier introuvable : {path}")

        wb = load_workbook(path, read_only=True, data_only=True)
        sheet = wb.active

        def normalized_row(cells):
            values = list(cells)
            values += [None] * (3 - len(values))
            return values[:3]

        if sheet.max_row < 1:
            raise CommandError("Fichier vide.")

        title, description, start_clue = normalized_row(
            [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        )
        if not title or not start_clue:
            raise CommandError("La première ligne doit contenir au minimum le titre et l'indice de départ.")

        with transaction.atomic():
            hunt = Hunt.objects.create(
                title=str(title).strip(),
                description=str(description or "").strip(),
                start_clue=str(start_clue).strip(),
            )

            created_steps = 0
            next_order = 1

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                order, name, clue_text = normalized_row(row)
                order = int(order) if order else next_order
                name = str(name or "").strip()
                clue_text = str(clue_text or "").strip()

                if not name or not clue_text:
                    raise CommandError(f"Ligne invalide (ordre {order}) : nom et indice sont requis.")

                Step.objects.create(
                    hunt=hunt,
                    order=order,
                    name=name,
                    clue_text=clue_text,
                )
                created_steps += 1
                next_order = max(next_order, order + 1)

        wb.close()
        self.stdout.write(self.style.SUCCESS(f"Chasse créée : {hunt.title} ({created_steps} étapes)"))
